"""
Parse Staff Strength Details Excel -> role_vacancy_clean.csv
Output format: prison, type, parent, designation, sanctioned, filled, vacancy

Prison names in CSV are normalized to match DB prison.jail names exactly.
jail_type in CSV matches DB jail_type values for unambiguous lookup.

Run:
    python scripts/parse_excel_vacancy.py "C:/Users/DELL/Downloads/Staff Strength Details (1).xlsx"
"""

import sys
import csv
import re
from openpyxl import load_workbook

# Map Excel role names -> canonical DEV role names (keys are lowercased/normalized)
ROLE_MAP = {
    'jailor':                           'Jailer',
    'asst. jailor':                     'Assistant Jailer',
    'assistant jailor':                 'Assistant Jailer',
    'deputy jailor':                    'Deputy Jailer',
    'first grade warder':               'Grade I Warder',
    'first grade warder (female)':      'Grade I Warder',
    'first grade warder (male)':        'Grade I Warder',
    'first grade warder women':         'Grade I Warder',
    'grade i warder':                   'Grade I Warder',
    'grade i warder women':             'Grade I Warder',
    'gr. i warder (female)':            'Grade I Warder',
    'gr. i warder (male)':              'Grade I Warder',
    'second grade warder':              'Grade II Warder',
    'second grade warder (female)':     'Grade II Warder',
    'second grade warder  (female)':    'Grade II Warder',
    'second grade warder (male)':       'Grade II Warder',
    'second grade warder women':        'Grade II Warder',
    'female second grade warder':       'Grade II Warder',
    'female second warder women':       'Grade II Warder',
    'grade ii warder':                  'Grade II Warder',
    'grade ii warder (women)':          'Grade II Warder',
    'grade ii warder women':            'Grade II Warder',
    'grade ii warder (vigilance)':      'Grade II Warder',
    'gr. ii warder (female)':           'Grade II Warder',
    'chief head warder':                'Chief Head Warder',
    'chief  head  warder':              'Chief Head Warder',
    'chief head warder (vigilance)':    'Chief Head Warder',
}

# Map Excel full prison names -> (db_name, db_jail_type)
# DB jail_type values: central_jail, district_jail, sub_jail, spw,
#   women_sub_jail, special_sub_jail, open_air_jail, farm_jail, transit_yard
PRISON_MAP = {
    # Central Prisons
    'Central Prison-I, Chennai':                    ('Chennai - I',                 'central_jail'),
    'Central Prison-II, Chennai':                   ('Chennai - II',                'central_jail'),
    'Central Prison, Vellore':                      ('Vellore',                     'central_jail'),
    'Central Prison, Cuddalore':                    ('Cuddalore',                   'central_jail'),
    'Central Prison, Salem':                        ('Salem',                       'central_jail'),
    'Central Prison, Coimbatore':                   ('Coimbatore',                  'central_jail'),
    'Central Prison, Tiruchirappalli':              ('Tiruchirappalli',             'central_jail'),
    'Central Prison, Madurai':                      ('Madurai',                     'central_jail'),
    'Central Prison, Palayamkottai':                ('Palayamkottai',               'central_jail'),
    # Special Prisons for Women (SPW)
    'Special Prison for Women, Chennai':            ('Chennai',                     'spw'),
    'Special Prison for Women, Vellore':            ('Vellore',                     'spw'),
    'Special Prison for Women, Coimbatore \xa0':    ('Coimbatore',                  'spw'),
    'Special Prison for Women, Coimbatore':         ('Coimbatore',                  'spw'),
    'Special Prison for Women, Tiruchirappalli':    ('Tiruchirappalli',             'spw'),
    'Special Prison for Women, Madurai':            ('Madurai',                     'spw'),
    # District Jails
    'District Jail, Chengalpattu':                  ('Chengalpattu',                'district_jail'),
    'DISTRICT JAIL VILLUPURAM':                     ('Villupuram',                  'district_jail'),
    'District Jail Dharmapuri':                     ('Dharmapuri',                  'district_jail'),
    'District Jail Tiruppur':                       ('Tiruppur',                    'district_jail'),
    'District Jail, Erode @ Gobichettipalayam':     ('Erode @ Gobichettipalayam',   'district_jail'),
    'District Jail, Pudukkottai':                   ('Pudukkottai',                 'district_jail'),
    'District Jail, Nagapattinam':                  ('Nagapattinam',                'district_jail'),
    'District Jail  Dindigul':                      ('Dindigul',                    'district_jail'),
    'District Jail – Ramanathapuram':           ('Ramanathapuram',              'district_jail'),
    'District Jail– Virudhunagar':              ('Virudhunagar',                'district_jail'),
    'District Jail– Theni':                     ('Theni',                       'district_jail'),
    'District Jail Thoothukudi @ Perurani':         ('Thoothukudi @ Perurani',      'district_jail'),
    'District Jail, Nagercoil':                     ('Kanniyakumari @ Nagercoil',   'district_jail'),
    # Special Sub Jails
    'Special Sub Jail, Poonnamallee':               ('Poonamallee (Men)',            'special_sub_jail'),
    'Special Sub Jail, Kokkirakulam':               ('Kokkirakulam (Women)',         'special_sub_jail'),
    'Special Sub Jail (Women), Salem\xa0':           ('Salem (Women)',                'special_sub_jail'),
    'Special Sub Jail (Women), Salem':               ('Salem (Women)',                'special_sub_jail'),
    'Sub Jail, Nanguneri':                          ('Nanguneri (Men)',              'special_sub_jail'),
    # Open Air Jails
    'Open Air Jail, Singanallur':                   ('Singanallur',                 'open_air_jail'),
    'Open Air Prison - Purasadai Udaippu':          ('Purasaraidaiudaippu',         'open_air_jail'),
    # Women Sub Jails
    'Sub Jail Women, Villupuram':                   ('Villupuram',                  'women_sub_jail'),
    'Women Sub Jail - Cuddalore':                   ('Cuddalore',                   'women_sub_jail'),
    'Women Sub Jail, Dharmapuri':                   ('Dharmapuri',                  'women_sub_jail'),
    'Female Sub Jail, Thiruvarur\xa0\xa0':           ('Thiruvarur',                  'women_sub_jail'),
    'Female Sub Jail, Thiruvarur':                  ('Thiruvarur',                  'women_sub_jail'),
    'Sub Jail Women, Nilakottai':                   ('Nilakottai',                  'women_sub_jail'),
    'Sub Jail Women Paramakudi':                    ('Paramakudi',                  'women_sub_jail'),
    'Sub Jail for Women, Thukalay':                 ('Thuckalay',                   'women_sub_jail'),
    'Female Jail, Perurani':                        ('Tiruppur (Annex)',             'women_sub_jail'),
    'District Jail, Chengalpattu- Women Annex':     ('Chengalpattu',                'district_jail'),
    'District Jail Tiruppur Women Annexure':        ('Female Annex, Tiruppur',      'sub_jail'),
    # Regular Sub Jails with name differences
    'Sub Jail, Kancheepuram':                       ('Kancheepuram',                'sub_jail'),
    'Sub Jail, Tiruttani':                          ('Tiruthani',                   'sub_jail'),
    'Sub Jail, Ponneri':                            ('Ponneri',                     'sub_jail'),
    'Sub Jail, Tiruvallur':                         ('Tiruvallur',                  'sub_jail'),
    'Sub Jail, Saidapet':                           ('Saidapet',                    'sub_jail'),
    'Sub Jail, Ambur':                              ('Ambur',                       'sub_jail'),
    'Sub Jail,  Arani (Closed)':                    ('CLOSED',                      'sub_jail'),
    'Sub Jail, Arakkonam':                          ('Arakkonam',                   'sub_jail'),
    'Sub Jail, Panruti':                            ('Panruti',                     'sub_jail'),
    'Sub Jail, Chidambaram':                        ('Chidambaram',                 'sub_jail'),
    'Sub Jail, Kallakurichi':                       ('Kallakurichi',                'sub_jail'),
    'Sub Jail, Tindivanam':                         ('Thindivanam',                 'sub_jail'),
    'Sub Jail, Virudhachalam':                      ('Virudhachalam',               'sub_jail'),
    'Sub Jail, Ulundurpet':                         ('Ulundurpet',                  'sub_jail'),
    'Sub Jail, Tirukkovilur':                       ('Thirukovilur',                'sub_jail'),
    'Sub Jail, Gingee':                             ('Gingee',                      'sub_jail'),
    'Sub Jail, Sankagiri':                          ('Sankagiri',                   'sub_jail'),
    'Sub Jail, Omalur':                             ('Omalur',                      'sub_jail'),
    'Sub Jail, Namakkal':                           ('Namakkal',                    'sub_jail'),
    'Sub Jail, Tiruchengodu':                       ('Thiruchengodu',               'sub_jail'),
    'Sub Jail, Harur':                              ('Harur',                       'sub_jail'),
    'Sub Jail, Krishnagiri':                        ('Krishnagiri',                 'sub_jail'),
    'Sub Jail, Hosur':                              ('Hosur',                       'sub_jail'),
    'Sub Jail, Uthangarai':                         ('Uthangarai',                  'sub_jail'),
    'Sub Jail, Erode':                              ('Erode',                       'sub_jail'),
    'Sub Jail, Bhavani':                            ('Bhavani',                     'sub_jail'),
    'Sub Jail, Sathiamangalam':                     ('Sathiamangalam',              'sub_jail'),
    'Sub Jail, Perundurai':                         ('Perundhurai',                 'sub_jail'),
    'Sub Jail, Coonoor':                            ('Coonoor',                     'sub_jail'),
    'Sub Jail, Gudalur':                            ('Gudalur',                     'sub_jail'),
    'Sub Jail, Ooty':                               ('Ooty',                        'sub_jail'),
    'Sub Jail, Palladam':                           ('Palladam',                    'sub_jail'),
    'Sub Jail, Tharapuram':                         ('Dharapuram',                  'sub_jail'),
    'Sub Jail, Avinasi':                            ('Avinashi',                    'sub_jail'),
    'Sub Jail, Udumalaipettai':                     ('Udumalaipettai',              'sub_jail'),
    'Sub Jail, Pollachi':                           ('Pollachi',                    'sub_jail'),
    'Sub Jail, Lalgudi':                            ('Lalgudi',                     'sub_jail'),
    'Sub Jail, Thuraiyur':                          ('Thuraiyur',                   'sub_jail'),
    'Sub Jail, Perambalur':                         ('Perambalur',                  'sub_jail'),
    'Sub Jail, Ariyalur':                           ('Ariyalur',                    'sub_jail'),
    'Sub Jail, Jeyakondam':                         ('Jeyankondam',                 'sub_jail'),
    'Sub Jail, Karur':                              ('Karur',                       'sub_jail'),
    'Sub Jail, Kulithalai':                         ('Kulithalai',                  'sub_jail'),
    'Sub Jail, Thirumayam':                         ('Thirumayam',                  'sub_jail'),
    'Sub Jail, Aranthangi':                         ('Aranthangi',                  'sub_jail'),
    'Sub Jail, Thanjavur':                          ('Thanjavur',                   'sub_jail'),
    'Sub Jail, Kumbakonam':                         ('Kumbakonam',                  'sub_jail'),
    'Sub Jail, Papanasam':                          ('Papanasam',                   'sub_jail'),
    'Sub Jail, Thiruvidaimaruthur':                 ('Thiruvidaimaruthur',          'sub_jail'),
    'Sub Jail, Mannarkudi':                         ('Mannargudi',                  'sub_jail'),
    'Sub Jail, Nannilam':                           ('Nannilam',                    'sub_jail'),
    'Sub Jail, Thiruthuraipoondi':                  ('Thiruthuraipoondi',           'sub_jail'),
    'Sub Jail, Mayiladuthurai':                     ('Mayiladuthurai',              'sub_jail'),
    'Sub Jail, Tharangampadi':                      ('Tharangambadi',               'sub_jail'),
    'Sub Jail, Sirkazhi':                           ('Sirkali',                     'sub_jail'),
    'Sub Jail, Aruppukottai':                       ('Aruppukottai',                'sub_jail'),
    'Sub Jail, Mudukulathur':                       ('Mudukulathur',                'sub_jail'),
    'Sub Jail, Periyakulam':                        ('Periakulam',                  'sub_jail'),
    'Sub Jail, Sivagangai':                         ('Sivagangai',                  'sub_jail'),
    'Sub Jail, Srivilliputhur':                     ('Srivilliputhur',              'sub_jail'),
    'Sub Jail, Thirumangalam':                      ('Thirumangalam',               'sub_jail'),
    'Sub Jail, Thiruppathur':                       ('Tiruppathur',                 'sub_jail'),
    'Sub Jail, Usilampatti':                        ('Usilampatti',                 'sub_jail'),
    'Sub Jail, Uthamapalayam':                      ('Uthamapalayam',               'sub_jail'),
    'Sub Jail, Vedasandur':                         ('Vedasandur',                  'sub_jail'),
    'Sub Jail, Ambasamudram':                       ('Ambasamuthiram',              'sub_jail'),
    'Sub Jail, Thenkasi':                           ('Tenkasi',                     'sub_jail'),
    'Sub Jail, Sankarankovil':                      ('Sankarankoil',                'sub_jail'),
    'Sub Jail, Srivaikundam':                       ('Srivaikundam',                'sub_jail'),
    'Sub Jail, Kuzhithurai':                        ('Kuzhithurai',                 'sub_jail'),
    'Sub Jail, Kovilpatti':                         ('Kovilpatti',                  'sub_jail'),
    # SICA (not in DB, will fail import gracefully)
    'State Institute of Correctional Administration (SICA), Tiruchirappalli': ('SICA Tiruchirappalli', 'sub_jail'),
}

# Sheets to skip (no prison-level data)
SKIP_SHEETS = {'Over All Vacancy ', 'Prison Hqrs', 'DIGs Range Office '}


def _resolve_prison(raw_name):
    """Return (db_name, db_type) for an Excel prison name."""
    stripped = raw_name.strip().rstrip('\xa0').strip()
    if stripped in PRISON_MAP:
        return PRISON_MAP[stripped]
    # Try stripping common prefixes to get city name
    for prefix in ['Sub Jail, ', 'Sub Jail,  ', 'Sub Jail ', 'District Jail, ',
                   'District Jail ', 'Central Prison, ', 'Special Sub Jail, ']:
        if stripped.startswith(prefix):
            city = stripped[len(prefix):].strip()
            if city:
                jail_type = 'central_jail' if 'Central Prison' in prefix else \
                            'district_jail' if 'District Jail' in prefix else \
                            'special_sub_jail' if 'Special Sub Jail' in prefix else 'sub_jail'
                return (city, jail_type)
    return (stripped, 'sub_jail')


def _normalize_role(raw):
    stripped = raw.strip()
    stripped = re.sub(r'\s*\(\s*\d+\s*M\s*[+\-]\s*\d+\s*F\s*\)', '', stripped).strip()
    key = re.sub(r'\s+', ' ', stripped.lower()).strip()
    return ROLE_MAP.get(key, stripped)


def _is_data_row(row):
    if row[0] is None or row[1] is None:
        return False
    try:
        float(row[0])
    except (TypeError, ValueError):
        return False
    if not isinstance(row[1], str):
        return False
    if row[2] is None:
        return False
    try:
        float(str(row[2]).replace('\xa0', '').strip())
    except (TypeError, ValueError):
        return False
    return True


def _is_prison_header(row):
    if row[0] is None:
        return False
    if not isinstance(row[0], str):
        return False
    if any(row[i] is not None for i in range(1, 5)):
        return False
    skip_patterns = ['sl.', 'sl. no', 'controlling sub jails', 'controlling jails',
                     'name of the post', '(hand made paper', ' sub jails', ' controlling']
    val = row[0].strip().lower()
    if any(p in val for p in skip_patterns):
        return False
    if len(val) < 4:
        return False
    return True


def _safe_int(val):
    if val is None:
        return 0
    try:
        return int(float(str(val).replace('\xa0', '').strip()))
    except (TypeError, ValueError):
        return 0


def parse_sheet(ws, sheet_name):
    records = []
    current_prison = None
    current_type = None

    for row in ws.iter_rows(values_only=True):
        row = list(row) + [None] * 5
        row = row[:5]

        if _is_prison_header(row):
            raw_name = row[0].strip()
            if '(closed)' in raw_name.lower() or 'closed' in raw_name.lower():
                current_prison = None
                current_type = None
                continue
            db_name, db_type = _resolve_prison(raw_name)
            if db_name == 'CLOSED':
                current_prison = None
                current_type = None
                continue
            current_prison = db_name
            current_type = db_type
            continue

        if current_prison and _is_data_row(row):
            role_raw = str(row[1]).strip()
            sanctioned = _safe_int(row[2])
            filled = _safe_int(row[3])
            vacant = _safe_int(row[4])
            role = _normalize_role(role_raw)
            if sanctioned == 0:
                continue
            # Only import the 6 executive staff roles tracked by the dashboard
            if role not in {'Jailer', 'Deputy Jailer', 'Assistant Jailer',
                            'Chief Head Warder', 'Grade I Warder', 'Grade II Warder'}:
                continue
            records.append({
                'prison': current_prison,
                'type': current_type,
                'parent': '',
                'designation': role,
                'sanctioned': sanctioned,
                'filled': filled,
                'vacancy': vacant,
            })

    return records


def main(xlsx_path, out_path='role_vacancy_clean.csv'):
    wb = load_workbook(xlsx_path)

    all_records = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        records = parse_sheet(ws, sheet_name)
        print(f'  {sheet_name}: {len(records)} rows')
        all_records.extend(records)

    # Aggregate by (prison, designation) to handle M+F rows for same prison
    agg = {}
    for r in all_records:
        k = (r['prison'], r['designation'])
        if k not in agg:
            agg[k] = {'type': r['type'], 'parent': r['parent'],
                      'sanctioned': 0, 'filled': 0, 'vacancy': 0}
        agg[k]['sanctioned'] += r['sanctioned']
        agg[k]['filled'] += r['filled']
        agg[k]['vacancy'] += r['vacancy']

    aggregated = []
    for (prison, designation), vals in agg.items():
        aggregated.append({
            'prison': prison,
            'type': vals['type'],
            'parent': vals['parent'],
            'designation': designation,
            'sanctioned': vals['sanctioned'],
            'filled': vals['filled'],
            'vacancy': vals['vacancy'],
        })

    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['prison', 'type', 'parent', 'designation', 'sanctioned', 'filled', 'vacancy'])
        writer.writeheader()
        writer.writerows(aggregated)

    all_records = aggregated
    print(f'\nTotal rows: {len(all_records)} -> {out_path}')

    key_roles = {'Jailer', 'Deputy Jailer', 'Assistant Jailer', 'Chief Head Warder', 'Grade I Warder', 'Grade II Warder'}
    from collections import defaultdict
    role_totals = defaultdict(lambda: [0, 0, 0])
    for r in all_records:
        if r['designation'] in key_roles:
            role_totals[r['designation']][0] += r['sanctioned']
            role_totals[r['designation']][1] += r['filled']
            role_totals[r['designation']][2] += r['vacancy']

    print('\n=== 6 Key Roles Summary ===')
    print(f"{'Role':<25} {'Sanctioned':>12} {'Filled':>8} {'Vacant':>8}")
    print('-' * 55)
    for role in ['Grade II Warder', 'Grade I Warder', 'Chief Head Warder', 'Assistant Jailer', 'Deputy Jailer', 'Jailer']:
        t = role_totals[role]
        print(f"{role:<25} {t[0]:>12} {t[1]:>8} {t[2]:>8}")
    totals = [sum(role_totals[r][i] for r in key_roles) for i in range(3)]
    print('-' * 55)
    print(f"{'TOTAL':<25} {totals[0]:>12} {totals[1]:>8} {totals[2]:>8}")


if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else 'Staff Strength Details (1).xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else 'role_vacancy_clean.csv'
    main(xlsx, out)
